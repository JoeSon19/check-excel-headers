import os
import sys
import openpyxl

FOLDER = os.path.dirname(os.path.abspath(__file__))

# Collect all Excel files in the folder
excel_files = sorted(
    f for f in os.listdir(FOLDER)
    if f.endswith((".xlsx", ".xls", ".xlsm")) and not f.startswith("~$")
)

if len(excel_files) < 2:
    print(f"Found {len(excel_files)} Excel file(s). Need at least 2 to compare.")
    sys.exit(1)

print(f"Found {len(excel_files)} Excel files.\n")

# Read header from the first file
first_file = excel_files[0]
wb = openpyxl.load_workbook(os.path.join(FOLDER, first_file), read_only=True)
ws = wb.active
reference_header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
wb.close()

print(f"Reference file: {first_file}")
print(f"Header: {reference_header}\n")

# Compare with every other file
for file_name in excel_files[1:]:
    wb = openpyxl.load_workbook(os.path.join(FOLDER, file_name), read_only=True)
    ws = wb.active
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    wb.close()

    if header != reference_header:
        print(f"MISMATCH: {file_name}")
        print(f"  Expected: {reference_header}")
        print(f"  Got:      {header}")
        sys.exit(1)
    else:
        print(f"  Matched: {file_name}")

print(f"\nAll {len(excel_files)} files have the same headers!")
